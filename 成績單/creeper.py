from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import re
import time
import sys
from openpyxl import Workbook
from selenium.webdriver.chrome.options import Options
LOGIN_URL = 'https://web085004.adm.ncyu.edu.tw/NewSite/login.aspx?Language=zh-TW'

def creep():
    if len(sys.argv) != 3:
        print("invalid format,please type like python creeper.py 'your_username' 'your_password'")
        return
    account = sys.argv[1]
    password = sys.argv[2]
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    driver = webdriver.Chrome(options=chrome_options)
    driver.get(LOGIN_URL)  # 進入登入網站
    act = driver.find_element_by_id("TbxAccountId")
    act.send_keys(account)
    pwd = driver.find_element_by_id("TbxPassword")
    pwd.send_keys(password)  # 輸入帳號密碼
    submit = driver.find_element_by_name("BtnPreLogin")
    submit.click()  # 登入
    wait = WebDriverWait(driver, 10)
    wait.until(lambda driver: driver.current_url != 'https://web085004.adm.ncyu.edu.tw/NewSite/login.aspx?Language=zh-TW')  # 等待跳轉頁面
    mode = driver.find_element_by_id("BtnMode")
    mode.click()
    button = driver.find_element_by_xpath("/html/body/div[4]/div[3]/button[2]")
    button.click()
    wait = WebDriverWait(driver, 10)
    wait.until(lambda driver: driver.current_url != 'https://web085004.adm.ncyu.edu.tw/NewSite/Index1.aspx')  # 等待跳轉頁面
    application = driver.find_element_by_link_text('學期成績查詢')
    application.click()  # 按學期成績查詢
    driver.switch_to_window(driver.window_handles[-1])
    button = driver.find_element_by_id("btnOK")
    button.click()
    time.sleep(0.2)
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    table = soup.find_all(id=re.compile('GridView1_ct*'))
    for i in range(len(table)):
        table[i] = table[i].text
    output = []
    for i in range(0, len(table), 5):
        output.append([table[i], table[i+1], table[i+2], table[i+3], table[i+4]])
    wb = Workbook()
    sheet = wb['Sheet']
    col_idx = [1, 2, 3, 4, 5, 6]
    for i in range(1, len(output)+1):
        for j in range(1, len(output[i-1])+1):
            if j == len(output[i-1])-1:
                sheet.cell(row=i, column=col_idx[j - 1], value=int(output[i - 1][j - 1]))
            else:
                sheet.cell(row=i, column=col_idx[j-1], value=output[i - 1][j - 1])

    wb.save('grade.xlsx')
    driver.close()
    #page = str(driver.page_source)
    #webpid = page[5006:5070]  # 無包含單引號
    #driver.execute_script("$.post(\"OpenLog.aspx\",{'WebPid1': '" + webpid + "', 'Program': 'stusco_630'}, function () { });")
    #referurl = "Refer.aspx?action=/grade_net/StuSco_630.aspx&amp;method=post&amp;WebPid1=" + webpid + "&amp;Language=zh-TW"
    #driver.execute_script("var $ApplicationFrame = $('#application-frame-main');$ApplicationFrame.attr('src', '" + referurl + "');")
creep()